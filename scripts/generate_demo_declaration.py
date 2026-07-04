from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "docs" / "declarations"
OUT_DIR.mkdir(parents=True, exist_ok=True)

generated_at = datetime.now().strftime("%Y%m%d-%H%M%S")
xlsx_path = OUT_DIR / f"dn-house-ke-khai-demo-{generated_at}.xlsx"
note_path = OUT_DIR / f"dn-house-ke-khai-demo-{generated_at}.md"

orders = [
    {
        "row": 3,
        "created_at": "04/07/2026 12:00:59",
        "order_no": "SYNC-120057-001",
        "customer_name": "Codex Test 1",
        "service_name": "Test POS sync",
        "quantity": 1,
        "unit_price": 7000,
        "subtotal": 7000,
        "discount": 0,
        "final_total": 7000,
        "status": "Da xoa",
    },
    {
        "row": 4,
        "created_at": "04/07/2026 12:01:02",
        "order_no": "SYNC-120057-002",
        "customer_name": "Codex Test 2",
        "service_name": "Test POS sync",
        "quantity": 2,
        "unit_price": 7000,
        "subtotal": 14000,
        "discount": 0,
        "final_total": 14000,
        "status": "",
    },
    {
        "row": 5,
        "created_at": "04/07/2026 12:01:04",
        "order_no": "SYNC-120057-003",
        "customer_name": "Codex Test 3",
        "service_name": "Test POS sync",
        "quantity": 3,
        "unit_price": 7000,
        "subtotal": 21000,
        "discount": 0,
        "final_total": 21000,
        "status": "Da xoa",
    },
    {
        "row": 6,
        "created_at": "04/07/2026 12:01:08",
        "order_no": "SYNC-120057-004",
        "customer_name": "Codex Test 4",
        "service_name": "Test POS sync",
        "quantity": 4,
        "unit_price": 7000,
        "subtotal": 28000,
        "discount": 0,
        "final_total": 28000,
        "status": "",
    },
    {
        "row": 7,
        "created_at": "04/07/2026 12:01:11",
        "order_no": "SYNC-120057-005",
        "customer_name": "Codex Test 5",
        "service_name": "Test POS sync",
        "quantity": 5,
        "unit_price": 7000,
        "subtotal": 35000,
        "discount": 0,
        "final_total": 35000,
        "status": "Da xoa",
    },
]

active_orders = [order for order in orders if order["status"] != "Da xoa"]
deleted_orders = [order for order in orders if order["status"] == "Da xoa"]

revenue = sum(order["final_total"] for order in active_orders)
deleted_revenue = sum(order["final_total"] for order in deleted_orders)
vat_rate = 0.05
pit_rate = 0.02
vat_amount = round(revenue * vat_rate)
pit_amount = round(revenue * pit_rate)

wb = Workbook()
ws = wb.active
ws.title = "KE_KHAI_NHAP"

blue = "1D315B"
orange = "C83B0D"
light_blue = "EAF3FF"
light_orange = "FFF3E8"
gray = "EEF2F7"
thin = Side(style="thin", color="D9E2F1")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws["A1"] = "DN HOUSE - BAN KE KHAI THU NHAP/DOANH THU DEMO"
ws["A1"].font = Font(size=16, bold=True, color=blue)
ws.merge_cells("A1:J1")
ws["A2"] = "Ban demo tu don test POS. Khong dung nop thue truc tiep neu chua doi chieu mau chinh thuc."
ws["A2"].font = Font(italic=True, color="666666")
ws.merge_cells("A2:J2")

summary_rows = [
    ("Ho kinh doanh", "Ho Kinh Doanh Giat Say DN House"),
    ("Nhom nganh tam tinh", "Dich vu giat say/ve sinh giay"),
    ("Ky ke khai demo", "Ngay 04/07/2026"),
    ("So don trong tap test", len(orders)),
    ("So don hop le tinh doanh thu", len(active_orders)),
    ("So don da xoa/loai tru", len(deleted_orders)),
    ("Doanh thu tinh thue demo", revenue),
    ("Doanh thu da xoa khong tinh", deleted_revenue),
    ("GTGT tam tinh 5%", vat_amount),
    ("TNCN tam tinh 2%", pit_amount),
    ("Tong thue tam tinh", vat_amount + pit_amount),
]

start_row = 4
for index, (label, value) in enumerate(summary_rows, start_row):
    ws.cell(index, 1, label)
    ws.cell(index, 2, value)
    ws.cell(index, 1).font = Font(bold=True, color=blue)
    ws.cell(index, 1).fill = PatternFill("solid", fgColor=light_blue)
    ws.cell(index, 2).fill = PatternFill("solid", fgColor="FFFFFF")
    for col in range(1, 3):
        ws.cell(index, col).border = border
        ws.cell(index, col).alignment = Alignment(vertical="center")

money_rows = {start_row + 6, start_row + 7, start_row + 8, start_row + 9, start_row + 10}
for row in money_rows:
    ws.cell(row, 2).number_format = '#,##0 "d"'

table_row = start_row + len(summary_rows) + 3
headers = [
    "Dong sheet",
    "Ngay tao",
    "Ma don",
    "Khach",
    "Dich vu",
    "SL",
    "Don gia",
    "Tam tinh",
    "Giam gia",
    "Tong thu",
    "Trang thai xu ly",
]
for col, header in enumerate(headers, 1):
    cell = ws.cell(table_row, col, header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=blue)
    cell.border = border
    cell.alignment = Alignment(horizontal="center", vertical="center")

for r, order in enumerate(orders, table_row + 1):
    included = order["status"] != "Da xoa"
    values = [
        order["row"],
        order["created_at"],
        order["order_no"],
        order["customer_name"],
        order["service_name"],
        order["quantity"],
        order["unit_price"],
        order["subtotal"],
        order["discount"],
        order["final_total"],
        "Tinh doanh thu" if included else "Da xoa - loai tru",
    ]
    for col, value in enumerate(values, 1):
        cell = ws.cell(r, col, value)
        cell.border = border
        cell.alignment = Alignment(vertical="center")
        if not included:
            cell.fill = PatternFill("solid", fgColor=gray)
        elif col == 11:
            cell.fill = PatternFill("solid", fgColor=light_orange)
            cell.font = Font(bold=True, color=orange)
    for col in [7, 8, 9, 10]:
        ws.cell(r, col).number_format = '#,##0 "d"'

notes_row = table_row + len(orders) + 3
notes = [
    "Ghi chu noi bo:",
    "- File nay chi la ban nhap tu don test POS, khong phai to khai nop thue chinh thuc.",
    "- Don co trang thai Da xoa duoc giu lai de doi chieu nhung khong tinh vao doanh thu.",
    "- Thue suat 5% GTGT va 2% TNCN chi la gia dinh demo cho nhom dich vu; can xac nhan lai voi co quan thue khi nop that.",
    "- Neu doanh thu nam thuoc nguong khong phai nop thue theo quy dinh, so thue thuc te co the bang 0.",
]
for i, text in enumerate(notes, notes_row):
    ws.cell(i, 1, text)
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=10)
    ws.cell(i, 1).alignment = Alignment(wrap_text=True)
    ws.cell(i, 1).font = Font(bold=i == notes_row, color=blue if i == notes_row else "333333")

widths = [12, 20, 22, 18, 20, 8, 14, 14, 14, 14, 20]
for col, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

ws.freeze_panes = f"A{table_row + 1}"

wb.save(xlsx_path)

note_path.write_text(
    "\n".join(
        [
            "# DN House - ban ke khai demo",
            "",
            f"- File Excel: `{xlsx_path.name}`",
            "- Nguon du lieu: Google Sheet `DON_HANG`, tap don test `SYNC-120057`.",
            "- Cach xu ly: don co trang thai `Da xoa` duoc giu lai de doi chieu, khong tinh vao doanh thu.",
            f"- Doanh thu hop le demo: {revenue:,} d".replace(",", "."),
            f"- GTGT tam tinh 5%: {vat_amount:,} d".replace(",", "."),
            f"- TNCN tam tinh 2%: {pit_amount:,} d".replace(",", "."),
            f"- Tong thue tam tinh: {vat_amount + pit_amount:,} d".replace(",", "."),
            "",
            "Luu y: day la ban demo noi bo, chua phai to khai nop thue chinh thuc.",
        ]
    ),
    encoding="utf-8",
)

print(xlsx_path)
print(note_path)
